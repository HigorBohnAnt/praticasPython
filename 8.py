import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

data = {
    "Jogador": ["Alpha", "Bravo", "Charlie", "Delta"],
    "Pontuação": [1500, 1200, 1800, 900],
    "Abates": [50, 40, 60, 30],
    "Mortes": [2, 3, 1, 4],
    "Missões Completas": [3, 2, 4, 1],
    "Precisão (%)": [85, 78, 92, 65],
    "Tempo em Missão (min)": [30, 28, 35, 20]
}

df = pd.DataFrame(data)

arquivo_excel = "helldivers_partida.xlsx"
df.to_excel(arquivo_excel, index=False)

wb = load_workbook(arquivo_excel)
sheet = wb.active

header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
header_font = Font(bold=True, color="000000")
alignment = Alignment(horizontal="center", vertical="center")

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = alignment

for column in sheet.columns:
    max_length = max(len(str(cell.value)) for cell in column if cell.value)
    adjusted_width = max_length + 2
    col_letter = column[0].column_letter
    sheet.column_dimensions[col_letter].width = adjusted_width

sheet.sheet_view.showGridLines = False

row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if row[0].row % 2 == 0:  
        for cell in row:
            cell.fill = row_fill

wb.save(arquivo_excel)

print(f"Planilha 'helldivers_partida.xlsx' criada e personalizada com sucesso!")

print(df)
